#!/usr/bin/env python3
"""
営業成績管理 Excel テンプレート生成スクリプト
使い方:
  python generate_excel.py          # 当月のテンプレートを生成
  python generate_excel.py 2026 3   # 指定年月のテンプレートを生成
"""

import sys
import datetime
import calendar

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl が必要です。以下を実行してください:")
    print("  pip install openpyxl")
    sys.exit(1)


# ===== スタイルヘルパー =====

def hex_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_border(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(cell, value, bg="3B82F6", fg="FFFFFF"):
    cell.value = value
    cell.font = Font(bold=True, color=fg, size=10)
    cell.fill = hex_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = make_border("2563EB")

def apply_label(cell, value):
    cell.value = value
    cell.font = Font(size=10, color="374151")
    cell.fill = hex_fill("F9FAFB")
    cell.alignment = Alignment(vertical="center", indent=1)
    cell.border = make_border()

def apply_value(cell, value=None, money=False, pct=False, bold=False, editable=False):
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, size=10, color="111827")
    cell.fill = hex_fill("FFFBEB" if editable else "FFFFFF")
    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    cell.border = make_border()
    if money:
        cell.number_format = "¥#,##0"
    elif pct:
        cell.number_format = '0.0"%"'

def section_header(ws, row, text, col_end=2):
    ws.merge_cells(f"A{row}:{get_column_letter(col_end)}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = hex_fill("1D4ED8")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


# ===== メイン生成 =====

def create_template(year=None, month=None):
    today = datetime.date.today()
    year  = year  or today.year
    month = month or today.month
    days  = calendar.monthrange(year, month)[1]
    dates = [datetime.date(year, month, d) for d in range(1, days + 1)]
    output = f"eigyo_{year}-{month:02d}.xlsx"

    wb = openpyxl.Workbook()

    # =========================================================
    # シート1: 成約明細
    # =========================================================
    ws_c = wb.active
    ws_c.title = "成約明細"
    ws_c.freeze_panes = "A2"

    headers_c = ["日付", "売上", "買取金額", "粗利", "備考"]
    widths_c  = [12, 15, 15, 15, 32]
    for i, (h, w) in enumerate(zip(headers_c, widths_c), 1):
        apply_header(ws_c.cell(1, i), h)
        ws_c.column_dimensions[get_column_letter(i)].width = w
    ws_c.row_dimensions[1].height = 28

    # 100行分のデータ領域
    for r in range(2, 102):
        ws_c.cell(r, 1).number_format = "YYYY/MM/DD"
        ws_c.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws_c.cell(r, 1).border = make_border()

        apply_value(ws_c.cell(r, 2), money=True, editable=True)
        apply_value(ws_c.cell(r, 3), money=True, editable=True)

        # 粗利 = 売上 - 買取（自動計算）
        gc = ws_c.cell(r, 4, value=f"=IF(B{r}=\"\",\"\",B{r}-C{r})")
        gc.number_format = "¥#,##0"
        gc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        gc.border = make_border()

        ws_c.cell(r, 5).alignment = Alignment(vertical="center")
        ws_c.cell(r, 5).border = make_border()

    # =========================================================
    # シート2: 日次集計
    # =========================================================
    ws_d = wb.create_sheet("日次集計")
    ws_d.freeze_panes = "A2"

    headers_d = ["日付", "不成約数", "不成約備考", "顧客数",
                 "成約数", "売上合計", "買取合計", "粗利合計", "対応数"]
    widths_d  = [12, 10, 28, 10, 10, 15, 15, 15, 10]
    for i, (h, w) in enumerate(zip(headers_d, widths_d), 1):
        apply_header(ws_d.cell(1, i), h)
        ws_d.column_dimensions[get_column_letter(i)].width = w
    ws_d.row_dimensions[1].height = 28

    # 累計列（チャート用、非表示にしてもOK）
    ws_d.cell(1, 11, "累計売上").font = Font(bold=True, size=9, color="9CA3AF")
    ws_d.cell(1, 12, "累計粗利").font = Font(bold=True, size=9, color="9CA3AF")
    ws_d.column_dimensions["K"].width = 14
    ws_d.column_dimensions["L"].width = 14

    for i, d in enumerate(dates, 2):
        r = i
        # 日付（入力済み）
        dc = ws_d.cell(r, 1, value=d)
        dc.number_format = "MM/DD(AAA)"
        dc.alignment = Alignment(horizontal="center", vertical="center")
        dc.border = make_border()
        dc.fill = hex_fill("F0F9FF")

        # 不成約数・不成約備考・顧客数（手入力）
        apply_value(ws_d.cell(r, 2), editable=True)
        nc = ws_d.cell(r, 3)
        nc.alignment = Alignment(vertical="center")
        nc.border = make_border()
        nc.fill = hex_fill("FFFBEB")
        apply_value(ws_d.cell(r, 4), editable=True)

        # 成約数 = COUNTIF(成約明細.A列, この日付)
        ec = ws_d.cell(r, 5, value=f"=COUNTIF(成約明細!$A:$A,A{r})")
        ec.alignment = Alignment(horizontal="center", vertical="center")
        ec.border = make_border()

        # 売上合計
        fc = ws_d.cell(r, 6, value=f"=SUMIF(成約明細!$A:$A,A{r},成約明細!$B:$B)")
        fc.number_format = "¥#,##0"
        fc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        fc.border = make_border()

        # 買取合計
        gc = ws_d.cell(r, 7, value=f"=SUMIF(成約明細!$A:$A,A{r},成約明細!$C:$C)")
        gc.number_format = "¥#,##0"
        gc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        gc.border = make_border()

        # 粗利合計
        hc = ws_d.cell(r, 8, value=f"=F{r}-G{r}")
        hc.number_format = "¥#,##0"
        hc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        hc.border = make_border()

        # 対応数 = 成約数 + 不成約数
        ic = ws_d.cell(r, 9, value=f"=E{r}+B{r}")
        ic.alignment = Alignment(horizontal="center", vertical="center")
        ic.border = make_border()

        # 累計（チャート用）
        if r == 2:
            ws_d.cell(r, 11, value=f"=F{r}")
            ws_d.cell(r, 12, value=f"=H{r}")
        else:
            ws_d.cell(r, 11, value=f"=K{r-1}+F{r}")
            ws_d.cell(r, 12, value=f"=L{r-1}+H{r}")
        ws_d.cell(r, 11).number_format = "¥#,##0"
        ws_d.cell(r, 12).number_format = "¥#,##0"

    last_row = 1 + days  # 日次集計の最終データ行

    # =========================================================
    # シート3: KPI
    # =========================================================
    ws_k = wb.create_sheet("KPI")
    ws_k.column_dimensions["A"].width = 24
    ws_k.column_dimensions["B"].width = 20

    # タイトル
    ws_k.merge_cells("A1:B1")
    tc = ws_k["A1"]
    tc.value = f"{year}年{month}月　営業KPI"
    tc.font = Font(bold=True, size=14, color="1E3A5F")
    tc.fill = hex_fill("DBEAFE")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_k.row_dimensions[1].height = 40

    # 目標設定（手入力 = 黄色背景）
    section_header(ws_k, 2, "■ 目標設定（手入力）")
    ws_k.row_dimensions[3].height = 26
    ws_k.row_dimensions[4].height = 26
    apply_label(ws_k.cell(3, 1), "売上目標")
    apply_value(ws_k.cell(3, 2), value=0, money=True, editable=True)
    apply_label(ws_k.cell(4, 1), "粗利目標")
    apply_value(ws_k.cell(4, 2), value=0, money=True, editable=True)

    # 累計実績
    section_header(ws_k, 5, "■ 累計実績")
    rows_actual = [
        ("累計売上",     f"=SUM(日次集計!F2:F{last_row})", True,  False, True),
        ("累計買取金額", f"=SUM(日次集計!G2:G{last_row})", True,  False, False),
        ("累計粗利",     f"=SUM(日次集計!H2:H{last_row})", True,  False, True),
        ("粗利率",       f"=IF(B6=0,0,B8/B6*100)",         False, True,  False),
    ]
    for i, (label, formula, money, pct_, bold) in enumerate(rows_actual, 6):
        ws_k.row_dimensions[i].height = 24
        apply_label(ws_k.cell(i, 1), label)
        apply_value(ws_k.cell(i, 2), value=formula, money=money, pct=pct_, bold=bold)

    # 目標対比
    section_header(ws_k, 10, "■ 目標対比")
    rows_target = [
        ("売上達成率",       f"=IF(B3=0,0,B6/B3*100)",  False, True),
        ("売上目標まであと", f"=MAX(0,B3-B6)",           True,  False),
        ("粗利達成率",       f"=IF(B4=0,0,B8/B4*100)",  False, True),
        ("粗利目標まであと", f"=MAX(0,B4-B8)",           True,  False),
    ]
    for i, (label, formula, money, pct_) in enumerate(rows_target, 11):
        ws_k.row_dimensions[i].height = 24
        apply_label(ws_k.cell(i, 1), label)
        apply_value(ws_k.cell(i, 2), value=formula, money=money, pct=pct_)

    # 対応・成約
    section_header(ws_k, 15, "■ 対応・成約")
    rows_ops = [
        ("成約数",       f"=SUM(日次集計!E2:E{last_row})"),
        ("不成約数",     f"=SUM(日次集計!B2:B{last_row})"),
        ("顧客数",       f"=SUM(日次集計!D2:D{last_row})"),
        ("合計対応数",   f"=SUM(日次集計!I2:I{last_row})"),
        ("平均粗利単価", f"=IF(B16=0,0,B8/B16)"),
    ]
    for i, row_data in enumerate(rows_ops, 16):
        ws_k.row_dimensions[i].height = 24
        apply_label(ws_k.cell(i, 1), row_data[0])
        c = ws_k.cell(i, 2, value=row_data[1])
        apply_value(c, money=(i == 20))

    # =========================================================
    # シート4: グラフ
    # =========================================================
    ws_g = wb.create_sheet("グラフ")
    ws_g.sheet_view.showGridLines = False

    chart = LineChart()
    chart.title  = f"{year}年{month}月　売上・粗利推移（累計）"
    chart.style  = 2
    chart.height = 16
    chart.width  = 28
    chart.y_axis.title = "金額（円）"
    chart.x_axis.title = "日付"
    chart.y_axis.numFmt = "¥#,##0"

    sales_data = Reference(ws_d, min_col=11, min_row=1, max_row=last_row)
    gross_data = Reference(ws_d, min_col=12, min_row=1, max_row=last_row)
    chart.add_data(sales_data, titles_from_data=True)
    chart.add_data(gross_data, titles_from_data=True)

    cats = Reference(ws_d, min_col=1, min_row=2, max_row=last_row)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.line.solidFill = "3B82F6"
    chart.series[0].graphicalProperties.line.width = 25000
    chart.series[1].graphicalProperties.line.solidFill = "F472B6"
    chart.series[1].graphicalProperties.line.width = 25000

    ws_g.add_chart(chart, "A1")

    # =========================================================
    # 凡例メモ（入力の説明）
    # =========================================================
    ws_k.cell(22, 1).value = "※ 黄色セルは手入力欄です"
    ws_k.cell(22, 1).font = Font(size=9, color="92400E")
    ws_k.cell(22, 1).fill = hex_fill("FEF3C7")
    ws_k.merge_cells("A22:B22")

    wb.save(output)
    print(f"✅  {output} を生成しました")
    print(f"    シート: 成約明細 / 日次集計 / KPI / グラフ")
    print(f"    ・成約明細: 日付・売上・買取を入力 → 粗利自動計算")
    print(f"    ・日次集計: 不成約数・備考・顧客数を入力 → 他は自動集計")
    print(f"    ・KPI: 目標を入力 → 達成率・差分が自動計算")


if __name__ == "__main__":
    year = month = None
    if len(sys.argv) == 3:
        year, month = int(sys.argv[1]), int(sys.argv[2])
    elif len(sys.argv) == 2:
        parts = sys.argv[1].split("-")
        year, month = int(parts[0]), int(parts[1])
    create_template(year=year, month=month)
